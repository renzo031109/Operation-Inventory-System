from django.shortcuts import render, redirect, get_object_or_404
from django.core.serializers.json import DjangoJSONEncoder
import json
from .forms import ClinicRecordFormSet, MedicineRecordFormSet, NewMedicineRecordFormSet
from .models import Clinic_Record, Medicine, MedCode, Location, MedicineMovement
from django.contrib import messages
from .filters import ClinicRecordFilter, MedicineFilter
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required

#for export excel imports
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import *
from urllib.parse import quote

from django.http import JsonResponse
import json


@login_required
def success(request):
    return render(request, 'clinic/success.html')


@login_required
def clinic_record_steps(request):
    insufficient_quantity = False  # Flag to track if any form fails quantity validation
    validation_errors = []  # Store error messages

    location_list = []
    employee_id_list = []
    last_name_list = []
    first_name_list = []
    gender_list = []
    company_list = []
    department_list = []
    illness_list = []
    amr_list = []
    medical_given_list = []
    note_list = []

    if request.method == 'POST':
        formset = ClinicRecordFormSet(request.POST)

        # Validate all forms first
        if formset.is_valid():
            records_to_process = []  # Temporary list to store valid forms and medicines
            for form in formset:
                if form.cleaned_data.get('medcode') and form.cleaned_data.get('quantity'):
                    get_medcode = form.cleaned_data.get('medcode')
                    get_quantity = form.cleaned_data.get('quantity')

                    try:
                        #code table
                        medicine_db = MedCode.objects.get(code=get_medcode)
                
                        # Check if quantity is sufficient
                        if medicine_db.quantity >= get_quantity:

                            #get the input of user
                            get_location = form.cleaned_data.get('location')
                            get_employee_id = form.cleaned_data.get('employee_id')
                            get_last_name = form.cleaned_data.get('last_name')
                            get_first_name = form.cleaned_data.get('first_name')
                            get_gender = form.cleaned_data.get('gender')
                            get_company = form.cleaned_data.get('company')
                            get_department = form.cleaned_data.get('department')
                            get_illness = form.cleaned_data.get('illness')
                            get_amr = form.cleaned_data.get('amr')
                            get_medical_given = form.cleaned_data.get('medical_given')
                            get_note = form.cleaned_data.get('note') 
                            # get_medicine = form.cleaned_data.get('medcode')
                            # get_quantity = form.cleaned_data.get('quantity')
                    
                            # Collect the valid form and associated medicine but do not deduct yet
                            records_to_process.append((form, medicine_db))
                            location_list.append(get_location)
                            employee_id_list.append(get_employee_id)
                            last_name_list.append(get_last_name)
                            first_name_list.append(get_first_name)
                            gender_list.append(get_gender)
                            company_list.append(get_company)
                            department_list.append(get_department)
                            illness_list.append(get_illness)
                            amr_list.append(get_amr)
                            medical_given_list.append(get_medical_given)
                            note_list.append(get_note)
                            
                        else:
                            insufficient_quantity = True
                            validation_errors.append(
                                f"Insufficient quantity for {medicine_db.medicine}. Available: {medicine_db.quantity}."
                            )
                    except Medicine.DoesNotExist:
                        validation_errors.append(f"Medicine {medicine_db.medicine} does not exist.")
                else:
                    validation_errors.append("Medicine and quantity are required for all forms.")

            # If any form fails validation, stop processing and show errors
            if insufficient_quantity or validation_errors:
                for error in validation_errors:
                    messages.error(request, error)
                return render(request, 'clinic/clinic_record_steps.html', {'formset': formset})

            # If all validations pass, process records and deduct quantities
            for form, medicine_db in records_to_process:
                # Deduct the quantity
                get_quantity = form.cleaned_data.get('quantity')
                medicine_db.quantity -= get_quantity

                    
                medicine_db.consumed += get_quantity
                medicine_db.save()

                #assign the first form value
                location = location_list[0]
                employee_id =  employee_id_list[0]
                last_name = last_name_list[0]
                first_name = first_name_list[0]
                gender = gender_list[0]
                company = company_list[0]
                department = department_list[0]
                illness = illness_list[0]
                amr = amr_list[0]
                medical_given = medical_given_list[0]
                note = note_list[0]

                # Save the form
                clinic_form = form.save(commit=False)
                clinic_form.medcode = medicine_db
                clinic_form.medicine = medicine_db.medicine
                clinic_form.quantity = -get_quantity  # Save negative quantity for record

                #assign the value to the fields db
                clinic_form.location = location
                clinic_form.employee_id =  employee_id
                clinic_form.last_name = last_name
                clinic_form.first_name = first_name
                clinic_form.gender = gender
                clinic_form.company = company
                clinic_form.department = department
                clinic_form.illness = illness
                clinic_form.amr = amr 
                clinic_form.medical_given = medical_given
                clinic_form.note = note

                clinic_form.save()


                #medicine movement logs
                medicine_movement = MedicineMovement(
                            code = medicine_db.code,
                            medicine = medicine_db.medicine,
                            quantity = -get_quantity,
                            note = "RELEASED",
                            location=location,
                            )
                medicine_movement.save()
 
                
            # Redirect to success page
            return redirect('success')

        else:
            messages.error(request, "Invalid input in one or more forms.")
    else:
        formset = ClinicRecordFormSet(queryset=Clinic_Record.objects.none())

    context = {'formset': formset}
    return render(request, 'clinic/clinic_record_steps.html', context)



@login_required
def add_medicine(request):

    location_list = []

    if request.method == 'POST':
        formset = MedicineRecordFormSet(request.POST)
        if formset.is_valid():
            for form in formset:
                
                # check if itemcode is selected
                if form.cleaned_data.get('medcode') and form.cleaned_data.get('quantity'):  
                    

                    #get the item qty from the form
                    add_location = form.cleaned_data.get('location')

                    #get the 1st input location to copy to next rows
                    location_list.append(add_location)

                    #get the item name from the form 
                    add_medicine = form.cleaned_data.get('medcode')

                    #get the item qty from the form
                    add_qty = form.cleaned_data.get('quantity')
             
                    medicine_soh = MedCode.objects.get(code=add_medicine)

                    #compute add soh
                    soh = int(medicine_soh.quantity) + int(add_qty)

                    #get the updated soh after add
                    medicine_soh.quantity = int(soh)
                    
                    #save tables
                    medicine_soh.save()

                    location = location_list[0]

                    #medicine movement logs
                    medicine_movement = MedicineMovement(
                            code = medicine_soh.code,
                            medicine = medicine_soh.medicine,
                            quantity = +add_qty,
                            note = "ADD",
                            location = location)
                    
                    medicine_movement.save()

                else:

                    messages.error(request, "Invalid Input. Form is incomplete.")
                    return redirect('add_medicine')
                
            messages.success(request, "You added stock successfully!")
                
            return redirect('medicine_report_details')
        
        else:
            messages.error(request, "Invalid Input!")

    else:
        formset = MedicineRecordFormSet(queryset=MedCode.objects.none())

    context = {'formset': formset}
    return render(request, 'clinic/add_medicine.html', context)


@login_required
def new_medicine(request):

    if request.method == 'POST':
        formset = NewMedicineRecordFormSet(request.POST)

        if formset.is_valid():
            for form in formset:
            
                # Get values from the form
                # Use cleaned_data for safety
                new_location = form.cleaned_data.get('location')
                new_medicine = form.cleaned_data.get('medicine')  
                new_medicine_qty = form.cleaned_data.get('quantity')
                new_demand = form.cleaned_data.get('demand')
                new_critical = form.cleaned_data.get('critical')

                #convert values of foreign key
                location_value = Location.objects.get(location=new_location)

                #formulate the itemcode
                concat = str(location_value) + " | " + new_medicine
       
                # Check if value exist in the database
                if MedCode.objects.filter(code__iexact=concat).exists():
                    messages.error(request, f"The value '{new_medicine}' already exists in the database. Please enter a different value.")
                    return render(request, 'clinic/new_medicine.html', {'formset': formset})
                
                # Assign code to Medcode table
                medcode = MedCode(
                    code=concat, 
                    location=location_value,
                    medicine=str(new_medicine), 
                    quantity=new_medicine_qty, 
                    demand=new_demand, 
                    critical=new_critical,
                    )
                medcode.save()
         
                # Check if fields are valid
                if new_medicine and new_medicine_qty:
                    # Create and save the new medicine entry
                    new_medicine_db = Medicine(
                        medcode=medcode,
                        medicine=str(new_medicine), 
                        quantity=new_medicine_qty, 
                        demand=new_demand, 
                        critical=new_critical, 
                        location=new_location)
                    new_medicine_db.save()


                    #medicine movement logs
                    medicine_movement = MedicineMovement(
                            code = concat,
                            medicine = str(new_medicine),
                            quantity = +new_medicine_qty,
                            note = "BEGINNING",
                            location = new_location)
                    medicine_movement.save()

                    messages.success(request, "You added stock successfully!")
                else:
                    # Add an error message for incomplete fields
                    messages.error(request, "Both medicine and quantity fields are required.")
                    return redirect('new_medicine')
                
            return redirect('medicine_report_details')
        else:
            messages.error(request, "Invalid Input!")

    else:
        formset = NewMedicineRecordFormSet(queryset=Medicine.objects.none())

    context = {'formset': formset}
    return render(request, 'clinic/new_medicine.html', context)


@login_required
def clinic_report_details(request):
    clinic = Clinic_Record.objects.all()

    clinic_record_filter = ClinicRecordFilter(request.GET, queryset=clinic)
    clinics = clinic_record_filter.qs
    item_count = clinics.count()

    #Prompt a message on the qty of the listed report
    if item_count > 0 :
        messages.info(request, f"Found '{item_count}' item(s) in the database")
    else:
        messages.info(request, f"Item not Found in the database ")


    #This will send the variable to filter views
    global filter_clinic_record_val
    def filter_clinic_record_val():
        return clinics


    context = {
        'item_count': item_count,
        'clinics': clinics,
        'clinic_record_filter': clinic_record_filter
        }
    return render(request, 'clinic/clinic_report_details.html', context)




@login_required
def medicine_movement_report(request):
    
    medicine = MedicineMovement.objects.all()

    # medicine_record_filter = MedicineFilter(request.GET, queryset=medicine)
    # medicines = medicine_record_filter.qs
    # item_count = medicines.count()

    # #Prompt a message on the qty of the listed report
    # if item_count > 0 :
    #     messages.info(request, f"Found '{item_count}' item(s) in the database")
    # else:
    #     messages.info(request, f"Item not Found in the database ")


    # #This will send the variable to filter views
    # global filter_medicine_record_val
    # def filter_medicine_record_val():
    #     return medicines


    context = {
        # 'item_count': item_count,
        'medicines': medicine,
        # 'clinic_record_filter': medicine_record_filter
        }
    
    return render(request, 'clinic/medicine_movement_report.html', context)




@login_required
def medicine_report_details(request):
    
    medicine = MedCode.objects.all()

    medicine_record_filter = MedicineFilter(request.GET, queryset=medicine)
    medicines = medicine_record_filter.qs
    item_count = medicines.count()

    #Prompt a message on the qty of the listed report
    if item_count > 0 :
        messages.info(request, f"Found '{item_count}' item(s) in the database")
    else:
        messages.info(request, f"Item not Found in the database ")


    #This will send the variable to filter views
    global filter_medicine_record_val
    def filter_medicine_record_val():
        return medicines


    context = {
        'item_count': item_count,
        'medicines': medicines,
        'clinic_record_filter': medicine_record_filter
        }
    return render(request, 'clinic/medicine_report.html', context)



@login_required
def clinic_export_excel_summary(request):

    #Export excel function
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="CLINICAL LOGS.xlsx"'

    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))


    # Declare Workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.merge_cells('A1:L1')

    first_cell = worksheet['A1']
    first_cell.value = "SHORE360 CLINICAL LOGS"
    first_cell.font = Font(bold=True)
    first_cell.alignment = Alignment(horizontal="center", vertical="center")


    worksheet.title = "CLINICAL LOGS"

    # Add headers
    headers =   [
                'LOCATION',	
                'EMPLOYEE ID',
                'NAME',
                'GENDER',
                'COMPANY',
                'DEPARTMENT/CLIENT',
                'CHIEF COMPLAINT',
                'BODY SYSTEM',
                'MEDICINE',
                'QUANTITY',
                'DATE'
                ]
    row_num = 2


    for col_num, column_title in enumerate(headers, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill("solid", fgColor="CFE2FF")
        cell.font = Font(bold=True, color="0B5ED7")
        cell.border = thin_border


    # Add data from the model
    # items = ItemBase.objects.all()
    # check if it was filtered if not set default to all
    if filter_clinic_record_val():
        clinics = filter_clinic_record_val()
        print("With filter value")
    else:
        clinics = Clinic_Record.objects.all()
        print("filter no value")

    for item in clinics:
        
        #convert object fields to string

        location = str(item.location)
        employee = str(item.employee_id)
        gender = str(item.gender)
        name = str(item.first_name+ ' ' + item.last_name)
        gender = str(item.gender)
        company = str(item.company)
        department = str(item.department)
        illness = str(item.illness)
        amr = str(item.amr)
        medicine = str(item.medicine)
        quantity = str(item.quantity)
        date_added = datetime.strftime(item.date_added,'%m/%d/%Y %H:%M:%S')


        worksheet.append([
            location,
            employee,
            name,
            gender,
            company,
            department,
            illness,
            amr,
            medicine,
            quantity,
            date_added
        ])
    
    workbook.save(response)
    return response


@login_required
def medicine_export_excel_summary(request):

    #Export excel function
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="MEDICINE LOGS.xlsx"'

    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))


    # Declare Workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.merge_cells('A1:E1')

    first_cell = worksheet['A1']
    first_cell.value = "MEDICINE LOGS"
    first_cell.font = Font(bold=True)
    first_cell.alignment = Alignment(horizontal="center", vertical="center")


    worksheet.title = "MEDICINE LOGS"

    # Add headers
    headers =   [
                'MEDICINE',
                'STOCK ON HAND',
                'TOTAL CONSUMED',
                'DEMAND',
                'CRITICAL QTY'
                ]
    row_num = 2


    for col_num, column_title in enumerate(headers, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill("solid", fgColor="CFE2FF")
        cell.font = Font(bold=True, color="0B5ED7")
        cell.border = thin_border


    # Add data from the model
    # items = ItemBase.objects.all()
    # check if it was filtered if not set default to all
    if filter_medicine_record_val():
        medicines = filter_medicine_record_val()
        print("With filter value")
    else:
        medicines = MedCode.objects.all()
        print("filter no value")

    for item in medicines:
        
        #convert object fields to string

        medicine = str(item.medicine)
        quantity = str(item.quantity)
        consumed = str(item.consumed)
        demand = str(item.demand)
        critical = str(item.critical)

        worksheet.append([
            medicine,
            quantity,
            consumed,
            demand,
            critical
        ])
    
    workbook.save(response)
    return response



#This is connected to medcode ajax value / promise
def load_medcode_code(request, location_id):
    medcode = list(MedCode.objects.filter(location_id=location_id).values('id', 'code'))
    return JsonResponse({'medcode': medcode})
















