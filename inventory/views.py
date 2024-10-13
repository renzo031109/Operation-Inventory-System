from typing import Any
from django.forms import BaseModelForm
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from .models import Item, ItemBase, ItemCode, UOM, Site, TeamMember, Floor
from .forms import ItemNewForm, ItemModelFormSet, ItemModelFormSetAdd
from .filters import ItemFilter, ItemBaseFilter
from django.http import HttpResponse
# from django.core.paginator import Paginator

#for export excel imports
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import *
from urllib.parse import quote

from django.http import JsonResponse
import json


# # #set universal variable for user settings.
# user_department = "FINANCE"
# user_client = "SHORE360"


@login_required
def summary_item(request):
    items = ItemBase.objects.all()
    # item_count_total= items.count()

    itemFilter = ItemBaseFilter(request.GET, queryset=items)
    items = itemFilter.qs
    item_count = items.count()

    if item_count > 0 :
        messages.info(request, f"Found '{item_count}' item(s) in the database")
    else:
        messages.info(request, f"Item not Found in the database ")

    # #pagination show 50 items per page
    # paginator = Paginator(items, 25)
    # page_number = request.GET.get("page")
    # page_obj = paginator.get_page(page_number)

    global filter_itembase_val
    def filter_itembase_val():
        return items

    # # Compute Total Value
    # total_value = 0
    # for value in items:
    #     total_value += float(value.total_value)

    context = {
        'items': items,
        'item_count': item_count,
        'itemFilter': itemFilter,
        # 'total_value': total_value
        # 'page_obj':page_obj
        }
    return render(request, 'inventory/summary.html', context)


@login_required
def inventory_item(request):
    
    items = Item.objects.all()
    # item_count_total= items.count()
    
    itemFilter = ItemFilter(request.GET, queryset=items)
    items = itemFilter.qs
    
    item_count = items.count()
    
    if item_count > 0 :
        messages.info(request, f"Found '{item_count}' transaction in the database")
    else:
        messages.info(request, f"Item not Found in the database ")
    
    # #pagination show 50 items per page
    # paginator = Paginator(items, 25)
    # page_number = request.GET.get("page")
    # page_obj = paginator.get_page(page_number)

    #this will get the current filter in the report
    global filter_item_val
    def filter_item_val():
        return items
    
    
    itembase = ItemBase.objects.all()

    context = {
        'items': items, 
        'itembase': itembase,
        'item_count': item_count, 
        'itemFilter': itemFilter,
        # 'page_obj': page_obj
        }
    return render(request,'inventory/inventory.html', context)


@login_required
def delete_item(request, id):
    if request.method == 'POST':

        try:
            #get the selected value
            item = Item.objects.get(id=id)
            item_soh = ItemBase.objects.get(item_code=item.item_code)

            #return the quantity of the deleted item
            if int(item.quantity) > 0:
                #computation on positive value
                updated_soh = int(item_soh.soh) - int(item.quantity) 

                # # updated_total_price = updated_soh * item_soh.price
                # total = item.quantity * item_soh.price
                # item_val = item_soh.total_value - total

                print("posive")
            else:
                #computation on negative value
                updated_soh = int(item_soh.soh) - int(item.quantity)

                # updated_total_price = updated_soh * item_soh.price
                # total = item.quantity * item_soh.price
                # item_val = item_soh.total_value - total

                print("negative")

            # item_soh.soh = updated_soh
            # item_soh.total_price = updated_total_price
            # item_soh.total_value = item_val

            #update Total value and save
            item_soh.save()
            item.delete()
            messages.success(request, "Item is deleted successfully")
        except:
            messages.error(request, "Unable to delete this item, please try again later")
    return redirect('inventory_item')



def delete_itembase(request, item_code):
    if request.method == 'POST':
        item = ItemBase.objects.get(item_code=item_code)
        itemcode = ItemCode.objects.get(code=item)

        print(item)
        print(itemcode)

        item.delete()
        itemcode.delete()

        messages.success(request, f"{item.item_name} is deleted successfully")
    return redirect('summary_item')



@login_required
def new_item(request):
    #initialize user client and department static

    if request.method == 'POST':
        form = ItemNewForm(request.POST)

        if form.is_valid():
            #get the value of the form
            form_item_name = request.POST.get('item_name')
            form_item_brand = request.POST.get('brand_name')
            form_item_soh = request.POST.get('soh')
            form_item_uom = request.POST.get('uom')
            # form_item_price = request.POST.get('price')

            #convert UOM id to values of foreign key
            uom_value = UOM.objects.get(id=form_item_uom)
         
            #using try-except method in case of null value
            try:
                record_name = ItemBase.objects.filter(item_name=form_item_name, brand_name=form_item_brand)

                for record in record_name:
                    if record.item_name.upper() == form_item_name.upper() and record.brand_name.upper() == form_item_brand.upper():
                        messages.error(request, "Item already exist!")
                        return redirect('new_item')
                       
            except:
                record_name = None

            #assign default value to remarks
            concat = form_item_name + " | " + form_item_brand
            itemcode = ItemCode(code=concat)

            #Assign form to a variable
            itemNewForm = form.save(commit=False)

            #define user for the staffname
            user = request.user
            # client = Client.objects.get(client=user_client)
            # department=Department.objects.get(department=user_department)


            #get the current user
            try: 
                user = request.user
                member = TeamMember.objects.get(member=user)
            except:
                member = TeamMember.objects.get(member="ADMIN")
           
            print(user)

            #copy newitem to Item Transaction
            itemTransaction = Item(
                                    item_code=itemcode, 
                                    item_name=form_item_name, 
                                    brand_name=form_item_brand, 
                                    quantity=form_item_soh, 
                                    remarks="BEGINNING",
                                    uom=uom_value, 
                                    # price=form_item_price,
                                    member=member,
                                    # client_name=client,
                                    # department_name=department
                                    )
            

            #assign generated code value to itemcode 
            itemNewForm.item_code = concat

            # #compute total value
            # total_value = float(form_item_price) * int(form_item_soh)

            # #beginning balance
            # itemNewForm.total_price = total_value
            # itemNewForm.total_value = total_value

            # #get item model and assign value to item_value
            # items = Item.objects.get(item_code=concat)
            # items.item_value = total_value

            try:
                #save tables if no error found
                itemNewForm.save()
                itemcode.save()
                itemTransaction.save()

                messages.success(request, "New Item added successfully!")
                return redirect('summary_item')
            except:
                messages.error(request, "Invalid Input")

    else:
        form = ItemNewForm()

    context = {'form': form}
    return render(request, 'inventory/new_item.html', context)


@login_required
def add_item(request):

    if request.method == 'POST':
        formset = ItemModelFormSetAdd(request.POST)
        if formset.is_valid():
            for form in formset:
                
                # check if itemcode is selected
                if form.cleaned_data.get('item_code') and form.cleaned_data.get('quantity'):  
                    
                    #get the item name from the form 
                    add_item_code = form.cleaned_data.get('item_code')
                    #get the item qty from the form
                    add_item_qty = form.cleaned_data.get('quantity')
                    # #get the price from the form
                    # add_item_price = form.cleaned_data.get('price')
                    
                    try:
                        #get the item SOH from model table
                        item_soh = ItemBase.objects.get(item_code=add_item_code)

                        # #assign value to client and department of authenticated user
                        # client = Client.objects.get(client=user_client)
                        # department = Department.objects.get(department=user_department)

    
                        #compute add soh
                        soh = int(item_soh.soh) + int(add_item_qty)
                        #get the updated soh after add
                        item_soh.soh = int(soh)
                        
                        # #compute item price
                        # added_item_amount = add_item_qty * add_item_price
                        
                        # #update Total value and price
                        # grand_total_value = item_soh.total_value + added_item_amount

                        # item_soh.total_value = grand_total_value
                        # item_soh.price = add_item_price
                        # item_soh.total_price = soh * item_soh.price

                        #save tables
                        item_soh.save()
                    except:
                        item_soh = 0
                        soh = int(item_soh) + int(add_item_qty)

                    #assign default values  
                    itemAddForm = form.save(commit=False)    
                    itemAddForm.remarks = "IN"
                    itemAddForm.item_name = item_soh.item_name
                    itemAddForm.brand_name = item_soh.brand_name
                    itemAddForm.uom = item_soh.uom
                    # itemAddForm.item_value = added_item_amount

                    #get the current user
                    try:
                        user = request.user
                        member = TeamMember.objects.get(member=user)
                        itemAddForm.member = member
                    except:
                        member = TeamMember.objects.get(member="ADMIN")
                        itemAddForm.member = member
                    # itemAddForm.client_name = client
                    # itemAddForm.department_name = department
  
                    itemAddForm.save()          
                
                else:

                    messages.error(request, "Invalid Input. Form is incomplete.")
                    return redirect('add_item')
                
            messages.success(request, "You added stock successfully!")
                
            return redirect('inventory_item')
        else:
            messages.error(request, "Invalid Input!")

    else:
        formset = ItemModelFormSetAdd(queryset=Item.objects.none())

    context = {'formset': formset}
    return render(request, 'inventory/add_item.html', context)


@login_required
def get_item(request):

    ##Initiate a list variable for the input select fields
    # staff_name_list = []
    # client_name_list = []
    # department_name_list = []
    member_name_list = []
    site_name_list = []
    floor_name_list = []
    purpose_name_list = []

    #list for validation checking
    item_error_list = []
    item_success_list = []

    if request.method == 'POST':
        formset = ItemModelFormSet(request.POST)
        if formset.is_valid():
            for form in formset:

                # only save if name is present
                if form.cleaned_data.get('item_code') and form.cleaned_data.get('quantity'): 
                                   
                    #get the item name from the form 
                    get_item_code = form.cleaned_data.get('item_code')

                    #get the item qty from the form
                    get_qty = form.cleaned_data.get('quantity')

                    #get the item SOH from model table       
                    item_soh = ItemBase.objects.get(item_code=get_item_code)


                    # # get the staff name values
                    # get_firstName = form.cleaned_data.get('firstName')
                    # get_middleName = form.cleaned_data.get('middleName')
                    # get_lastName = form.cleaned_data.get('lastName')

                    # # formatting of Full Name
                    # get_staff_name = f"{get_firstName} {get_middleName} {get_lastName}"
                    # get_client_name = form.cleaned_data.get('client_name')
                    # get_department_name = form.cleaned_data.get('department_name')
                    get_member_name = form.cleaned_data.get('member')
                    get_site_name = form.cleaned_data.get('site')
                    get_floor_name = form.cleaned_data.get('floor')
                    get_purpose_name = form.cleaned_data.get('purpose')


                    # #populate the list from the user input
                    # staff_name_list.append(get_staff_name)
                    # client_name_list.append(get_client_name)
                    # department_name_list.append(get_department_name)
                    member_name_list.append(get_member_name)
                    site_name_list.append(get_site_name)
                    floor_name_list.append(get_floor_name)
                    purpose_name_list.append(get_purpose_name)
                    
                    # #get the first value of the form
                    # staff_name = staff_name_list[0]
                    # client_name = client_name_list[0]
                    # department_name = department_name_list[0]
                    member_name = member_name_list[0]
                    site_name = site_name_list[0]
                    floor_name = floor_name_list[0]
                    purpose_name = purpose_name_list[0]

                    if item_soh.soh < get_qty:
                        messages.error(request, f"Ooops, Your available stock for '{item_soh.item_name}' is only '{item_soh.soh}")
                        item_error_list.append(item_soh.item_name)
                    else:
                        item_success_list.append(item_soh.item_name)
                        #compute add soh
                        soh = int(item_soh.soh) - int(get_qty)
                        #get the updated soh after add
                        item_soh.soh = int(soh)
                        #update Total value
                        # total = get_qty * item_soh.price
                        # total_val = item_soh.total_value - total
                        # item_soh.total_price = soh * item_soh.price
                        # item_soh.total_value = total_val
                        #save tables
                        item_soh.save()

                        #Convert qty value to negative for get
                        qtyToNegative = (get_qty) * -1

                        #assign default value to remarks 
                        itemGetForm = form.save(commit=False)    
                        itemGetForm.remarks = "OUT"
                        itemGetForm.item_name = item_soh.item_name
                        itemGetForm.brand_name = item_soh.brand_name
                        # itemGetForm.price = item_soh.price
                        itemGetForm.quantity = qtyToNegative
                        itemGetForm.uom = item_soh.uom
                        # itemGetForm.staff_name = staff_name
                        # itemGetForm.client_name = client_name
                        # itemGetForm.department_name = department_name
                        # itemGetForm.item_value = total
                        itemGetForm.member = member_name
                        itemGetForm.site = site_name
                        itemGetForm.floor = floor_name
                        itemGetForm.purpose = purpose_name

                        itemGetForm.save()
                else:
                    messages.error(request, "Invalid Input. Form is incomplete.")

            #assign length variables
            invalid_form = len(item_error_list)
            valid_form = len(item_success_list)

            # if values are valid send to submitted template
            if invalid_form == 0 and valid_form != 0:
                return redirect('submitted')

            # if 1 of the values are valid 
            elif invalid_form > 0 and valid_form > 0:
                messages.info(request, f"'{valid_form}' item(s) is/are submitted. Please re-input the item(s) with insufficient stock on hand.")
                return redirect('get_item')    
            # re input the items     
            else:
                return redirect('get_item')
            
        else:
            messages.error(request, "Invalid Input!")

    else:
        formset = ItemModelFormSet(queryset=Item.objects.none())

    context = {'formset': formset}
    return render(request, 'inventory/get_item.html', context)


@login_required
def submitted(request):
    return render(request, 'inventory/submitted.html')


@login_required
def export_excel_inventory(request):

    #Export excel function
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="DETAILED REPORT.xlsx"'

    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))


    # Declare Workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.merge_cells('A1:I1')

    first_cell = worksheet['A1']
    first_cell.value = "DETAILED REPORT"
    first_cell.font = Font(bold=True)
    first_cell.alignment = Alignment(horizontal="center", vertical="center")


    worksheet.title = "DETAILED REPORT"

    # Add headers
    headers =   [
                'ITEM NAME',	
                'BRAND NAME',
                'QUANTITY',	
                # 'PRICE',
                'UOM',	
                'DATE',
                'REMARKS',	
                'STAFF NAME',
                'SITE',
                'FLOOR',
                'PURPOSE'	
                ]
    row_num = 2


    for col_num, column_title in enumerate(headers, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill("solid", fgColor="CFE2FF")
        cell.font = Font(bold=True, color="0B5ED7")
        cell.border = thin_border


    # Add data from the model
    # items = Item.objects.all()
    # check if filtered items if not set default to all
    if filter_item_val():
        items = filter_item_val()
    else:
        items = Item.objects.all()

    for item in items:
        
        #convert object fields to string
        member_name = str(item.member)
        # client_name = str(item.client_name)
        # department_name = str(item.department_name)
        site_name = str(item.site)
        floor_name = str(item.floor)
        date_added = datetime.strftime(item.date_added,'%m/%d/%Y %H:%M:%S')

        worksheet.append([
            item.item_name,
            item.brand_name,
            item.quantity,
            # item.price,
            item.uom,
            date_added,
            item.remarks,
            member_name,
            site_name,
            floor_name,
            item.purpose
        ])
    
    workbook.save(response)
    return response


@login_required
def export_excel_summary(request):

    #Export excel function
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="INVENTORY REPORT.xlsx"'

    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))


    # Declare Workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.merge_cells('A1:G1')

    first_cell = worksheet['A1']
    first_cell.value = "INVENTORY REPORT"
    first_cell.font = Font(bold=True)
    first_cell.alignment = Alignment(horizontal="center", vertical="center")


    worksheet.title = "INVENTORY REPORT"

    # Add headers
    headers =   [
                'ITEM NAME',	
                'BRAND NAME',
                'UOM',	
                'SOH',
                # 'PRICE',	
                # 'TOTAL PRICE',
                # 'TOTAL BALANCE',
                'DATE'
                ]
    row_num = 2


    for col_num, column_title in enumerate(headers, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill("solid", fgColor="CFE2FF")
        cell.font = Font(bold=True, color="0B5ED7")
        cell.border = thin_border


    # Add data from the model
    # items = ItemBase.objects.all()
    # check if it was filtered if not set default to all
    if filter_itembase_val():
        items = filter_itembase_val()
        print("With filter value")
    else:
        items = ItemBase.objects.all()
        print("filter no value")

    for item in items:
        
        #convert object fields to string

        uom = str(item.uom)
        date_added = datetime.strftime(item.date_added,'%m/%d/%Y %H:%M:%S')

        worksheet.append([
            item.item_name,
            item.brand_name,
            uom,
            item.soh,
            # item.price,
            # item.total_price,
            # item.total_value,
            date_added,
        ])
    
    workbook.save(response)
    return response


def load_floor(request):
    data = json.loads(request.body)
    floor = Floor.objects.filter(site__id=data['user_id'])
    print(floor)
    return JsonResponse(list(floor.values('id','floor')), safe=False)
