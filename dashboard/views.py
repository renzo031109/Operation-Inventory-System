from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from inventory.models import Item, ItemBase
from clinic.models import Clinic_Record, MedCode
from django.http import HttpResponse
import datetime
from django.db.models import Q

from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl import Workbook
from openpyxl.styles import *

#set universal variable for value critical
none_value = 10

# Create your views here.
@login_required
def dashboard_view(request):

    #initialize critical value
    
    itembase = ItemBase.objects.all()
    item = Item.objects.all()
    clinic = Clinic_Record.objects.all()
    medicine = MedCode.objects.all()

    #Total Item Count
    item_count = ItemBase.objects.filter(site=1).count()
    medicine_count = MedCode.objects.filter(location=1).count()

    #Total Items with critical stocks
    critical_count = 0
    # critical_list = []
    for critical in itembase:
        
        #exclude critical value zero
        if critical.critical_value != 0:

            #condition when admin put a critical value
            if critical.critical_value is not None:
            
                if critical.soh <= critical.critical_value:      
                    critical_count += 1      
                    # critical_list.append(critical.id)

            #condition when critical value is empty
            else:   
                if critical.soh <= none_value:
                    critical_count += 1
                    # critical_list.append(critical.id)
 

                
    #Transaction today
    transaction_count = 0
    for transaction_date in item:
        if transaction_date.date_added.date() == datetime.datetime.now().date():
            transaction_count += 1


    #Count critical for medicine
    critical_medicine_count = 0
    for medcritical in medicine:
        if medcritical.quantity <= medcritical.critical:
            critical_medicine_count += 1


    # Clinic log today
    unique_users = set()  # A set to store unique user IDs
    for clinic_date in clinic:
        if clinic_date.date_added.date() == datetime.datetime.now().date():
            unique_users.add(clinic_date.employee_id)
    clinic_log_count = len(unique_users)  # Count the unique users


    context = {
        'item_count': item_count,
        'critical_count': critical_count,
        'none_value': none_value,
        'itembase': itembase,
        'transaction_count': transaction_count,
        'clinic_log_count': clinic_log_count,
        'medicine': medicine,
        'critical_medicine_count': critical_medicine_count,
        'medicine_count': medicine_count

    }

    return render(request, 'dashboard/dashboard_template.html', context)




def critical_stock_excel_export(request):
    #Export excel function
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="ITEM WITH CRITICAL STOCK.xlsx"'

    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

    # Declare Workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.merge_cells('A1:G1')

    first_cell = worksheet['A1']
    first_cell.value = "ITEM WITH CRITICAL STOCK"
    first_cell.font = Font(bold=True)
    first_cell.alignment = Alignment(horizontal="center", vertical="center")


    worksheet.title = "ITEM WITH CRITICAL STOCK"

    # Add headers
    headers =   [
                'site',
                'ITEM NAME',	
                'BRAND NAME',
                'UOM',	
                'SOH',	
                # 'PRICE',	
                'CRITICAL VALUE',
                'DEMAND'
                ]
    row_num = 2


    for col_num, column_title in enumerate(headers, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill("solid", fgColor="CFE2FF")
        cell.font = Font(bold=True, color="0B5ED7")
        cell.border = thin_border


    # Add data from the model
    items = ItemBase.objects.all()


    for item in items:
        
        if item.critical_value != 0:
        
            #condition when critical_value has value
            if item.critical_value > 0:
                if item.soh <= item.critical_value:
                    #convert object fields to string
                    uom = str(item.uom)
                    demand_item = str(item.demand_item)
                    site = str(item.site)

                    worksheet.append([
                    site,
                    item.item_name,
                    item.brand_name,
                    uom,
                    item.soh,
                    # item.price,
                    item.critical_value,
                    demand_item
                ])

            #condition when critical_value is null  
            else:
                if item.soh <= none_value :

                    #convert object fields to string
                    uom = str(item.uom)
                    worksheet.append([
                    item.item_name,
                    item.brand_name,
                    uom,
                    item.soh,
                    # item.price,
                    item.critical_value,
                    item.demand_item
                    ])
    
    workbook.save(response)
    return response



def critical_medicine_excel_export(request):
    #Export excel function
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="CLINIC CRITICAL STOCK.xlsx"'

    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

    # Declare Workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.merge_cells('A1:E1')

    first_cell = worksheet['A1']
    first_cell.value = "CLINIC CRITICAL STOCK"
    first_cell.font = Font(bold=True)
    first_cell.alignment = Alignment(horizontal="center", vertical="center")


    worksheet.title = "CLINIC CRITICAL STOCK"

    # Add headers
    headers =   [
                'LOCATION',
                'MEDICINE',	
                'SOH',		
                'CRITICAL VALUE',
                'DEMAND'
                ]
    row_num = 2


    for col_num, column_title in enumerate(headers, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill("solid", fgColor="CFE2FF")
        cell.font = Font(bold=True, color="0B5ED7")
        cell.border = thin_border


    # Add data from the model
    items = MedCode.objects.all()


    for item in items:
        
        if item.critical != 0:
        
            #condition when critical_value has value
            if item.critical > 0:
                if item.quantity <= item.critical:
                    #convert object fields to string
                    location = str(item.location)
                    demand = str(item.demand)

                    worksheet.append([
                    location,
                    item.medicine,
                    item.quantity,
                    item.critical,
                    demand
                ])

            #condition when critical_value is null  
            else:
                if item.quantity <= none_value :

                    #convert object fields to string
                    location = str(item.location)
                    demand = str(item.demand)

                    worksheet.append([
                    location,
                    item.medicine,
                    item.quantity,
                    item.critical,
                    demand

                    ])
    
    workbook.save(response)
    return response



